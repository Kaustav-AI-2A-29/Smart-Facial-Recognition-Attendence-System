"""
Face Recognition Attendance System
====================================
Requirements:
    pip install face_recognition opencv-python numpy

Dataset Structure:
    dataset/
        person_name/
            image1.jpg
            image2.jpg
"""

import face_recognition
import cv2
import numpy as np
import csv
import os
import time
import sys
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Add backend imports for database integration
sys.path.insert(0, os.path.abspath(os.path.dirname(__file__)))
from backend.database import db
from backend.attendance_service import record_attendance

# ─────────────────────────────────────────────
#  CONFIGURATION
# ─────────────────────────────────────────────
DATASET_DIR       = "dataset"       # folder containing sub-folders per person
ATTENDANCE_FILE   = "attendance.csv"
SCREENSHOTS_DIR   = "screenshots"   # folder for saving attendance photos
MATCH_THRESHOLD   = 0.48            # lower = stricter match (tightened for accuracy)
CONFIRM_FRAMES    = 10              # consecutive frames needed to confirm identity
PROCESS_EVERY_N   = 2               # process every Nth frame (performance)
DEBUG_MODE        = False           # Set to True only during development


# ─────────────────────────────────────────────
#  STEP 1 – LOAD DATASET & BUILD ENCODINGS
# ─────────────────────────────────────────────
def load_dataset(dataset_dir: str):
    """
    Walk dataset/ directory.
    Each sub-folder name  →  person name.
    Each image inside     →  face encoding.
    Returns (known_encodings, known_names).
    """
    known_encodings = []
    known_names     = []

    if not os.path.isdir(dataset_dir):
        print(f"[ERROR] Dataset folder '{dataset_dir}' not found.")
        return known_encodings, known_names

    # FIX 19: Use sorted() for deterministic ordering across platforms
    for person_name in sorted(os.listdir(dataset_dir)):
        person_dir = os.path.join(dataset_dir, person_name)
        if not os.path.isdir(person_dir):
            continue

        image_count = 0
        for filename in sorted(os.listdir(person_dir)):
            if not filename.lower().endswith((".jpg", ".jpeg", ".png")):
                continue

            image_path = os.path.join(person_dir, filename)

            # Use cv2 to load — handles JPEG, PNG, RGBA, grayscale, etc.
            bgr_image = cv2.imread(image_path)
            if bgr_image is None:
                print(f"  [WARN] Could not read image: {image_path}")
                continue

            # face_recognition expects RGB, not BGR
            rgb_image = cv2.cvtColor(bgr_image, cv2.COLOR_BGR2RGB)

            encodings = face_recognition.face_encodings(rgb_image)

            if encodings:
                known_encodings.append(encodings[0])
                known_names.append(person_name)
                image_count += 1
            else:
                print(f"  [WARN] No face found in: {image_path}")

        print(f"  Loaded {image_count} encoding(s) for: {person_name}")

    print(f"\n[INFO] Dataset ready – {len(known_encodings)} total encoding(s) "
          f"for {len(set(known_names))} person(s).\n")
    return known_encodings, known_names


# ─────────────────────────────────────────────
#  STEP 2 – ATTENDANCE FILE HELPERS
# ─────────────────────────────────────────────
def ensure_csv_header(filepath: str):
    """Write header row if file is new/empty."""
    write_header = not os.path.isfile(filepath) or os.path.getsize(filepath) == 0
    if write_header:
        with open(filepath, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(["Name", "Time (HH:MM:SS)", "Date (YYYY-MM-DD)"])


def load_todays_attendance(filepath: str) -> set:
    """Return a set of names already marked today."""
    marked = set()
    today  = datetime.now().strftime("%Y-%m-%d")

    if not os.path.isfile(filepath):
        return marked

    with open(filepath, newline="") as f:
        reader = csv.reader(f)
        next(reader, None)  # skip header row
        for row in reader:
            if len(row) >= 3 and row[2] == today:
                marked.add(row[0])
    return marked


def mark_attendance(name: str, filepath: str, marked_today: set):
    """Append a new attendance record if not already marked today."""
    if name in marked_today:
        return  # already recorded

    now   = datetime.now()
    today = now.strftime("%Y-%m-%d")
    time_str  = now.strftime("%H:%M:%S")

    with open(filepath, "a", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([name, time_str, today])

    marked_today.add(name)
    print(f"[ATTENDANCE] ✓ {name} marked at {time_str} on {today}")


def get_student_id_by_name(name: str) -> str:
    """
    Get student_id from database by name (case-insensitive).
    
    Args:
        name: Student name (from dataset folder name)
    
    Returns:
        student_id or empty string if not found
    """
    # Use LOWER() for case-insensitive matching
    result = db.execute_one(
        "SELECT student_id FROM students WHERE LOWER(name) = LOWER(?)",
        (name,)
    )
    return result["student_id"] if result else ""


def save_attendance_to_database(name: str, screenshot_path: str = None, confidence: float = 0.0):
    """
    Save attendance record to database using the attendance_service.
    
    Args:
        name: Student name (from face detection)
        screenshot_path: Path to screenshot file
        confidence: Face recognition confidence (0-100)
    """
    # FIX 1: Normalize to lowercase for consistency
    name = name.lower()
    student_id = get_student_id_by_name(name)
    if not student_id:
        print(f"[WARN] Could not find student_id for {name}")
        return
    
    try:
        # Use the standard record_attendance function from attendance_service
        record_attendance(
            student_id=student_id,
            screenshot_path=screenshot_path,
            confidence=confidence,  # Pass the actual confidence from identify_face
            liveness_passed=True,
            marked_by="system"
        )
        print(f"[DB] Attendance saved for {name} ({student_id}) with confidence {confidence:.1f}%")
    except Exception as e:
        print(f"[ERROR] Failed to save attendance to database: {e}")


# ─────────────────────────────────────────────
#  STEP 3 – FACE MATCHING
# ─────────────────────────────────────────────
def identify_face(face_encoding, known_encodings, known_names, threshold):
    """
    Compare face_encoding against all known encodings.
    Returns tuple of (name, confidence) where confidence is 0-100.
    """
    if not known_encodings:
        return "Unknown", 0.0

    distances = face_recognition.face_distance(known_encodings, face_encoding)
    best_idx  = int(np.argmin(distances))
    best_dist = distances[best_idx]

    # Convert distance to confidence percentage (lower distance = higher confidence)
    confidence = max(0.0, (1.0 - best_dist) * 100.0)

    # Debug: show best candidate and its distance (disabled by default)
    if DEBUG_MODE:
        print(f"[DEBUG] Best match: {known_names[best_idx]} | Distance: {best_dist:.3f} | Confidence: {confidence:.1f}%")

    # Strict matching – must beat threshold, otherwise treat as Unknown
    if best_dist < threshold:
        return known_names[best_idx], confidence
    else:
        return "Unknown", confidence


# ─────────────────────────────────────────────
#  STEP 4 – STABILITY TRACKER
# ─────────────────────────────────────────────
class StabilityTracker:
    """
    Tracks a detected name across frames.
    Identity is confirmed only when the same name appears
    for CONFIRM_FRAMES consecutive frames.
    """

    def __init__(self, required_frames: int):
        self.required_frames  = required_frames
        self.candidate        = None   # current candidate name
        self.streak           = 0      # consecutive frames for candidate
        self.confirmed        = None   # locked confirmed name
        self.confirmed_locked = False  # True once identity is confirmed

    def update(self, detected_name: str):
        """
        Feed the latest detected name.
        Returns (confirmed_name_or_None, display_name).
        confirmed_name_or_None  →  non-None only on the frame of confirmation
        display_name            →  what to show in the UI right now
        """
        # FIX 4: Do NOT short-circuit if confirmed_locked — allow re-detection
        if detected_name == self.candidate:
            self.streak += 1
        else:
            self.candidate = detected_name
            self.streak    = 1
            self.confirmed_locked = False  # unlock when face changes

        # Check confirmation threshold
        if self.streak >= self.required_frames and self.candidate != "Unknown":
            self.confirmed        = self.candidate
            self.confirmed_locked = True
            return self.confirmed, self.confirmed  # newly confirmed

        return None, self.candidate  # still building streak

    def reset(self):
        """Call this after attendance is marked to allow next detection."""
        self.candidate        = None
        self.streak           = 0
        self.confirmed        = None
        self.confirmed_locked = False


# ─────────────────────────────────────────────
#  STEP 5 – DRAW OVERLAY ON FRAME
# ─────────────────────────────────────────────
def draw_face_box(frame, top, right, bottom, left, name, confirmed: bool):
    """
    Green box  + name  →  confirmed identity
    Orange box + label →  unknown / still detecting
    """
    color = (0, 200, 0) if confirmed else (0, 140, 255)

    # Bounding box
    cv2.rectangle(frame, (left, top), (right, bottom), color, 2)

    # Label background
    label = name
    (tw, th), _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.7, 2)
    cv2.rectangle(frame, (left, bottom), (left + tw + 10, bottom + th + 12), color, -1)

    # Label text
    cv2.putText(frame, label, (left + 5, bottom + th + 5),
                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)


def draw_status_bar(frame, tracker: StabilityTracker, marked_today: set):
    """Draw a small status overlay at the top of the frame."""
    status_lines = [
        f"Marked today: {', '.join(sorted(marked_today)) if marked_today else 'None'}",
        "Press 'Q' to quit  |  Press 'R' to reset attendance",
    ]
    for i, line in enumerate(status_lines):
        cv2.putText(frame, line, (10, 25 + i * 22),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.55, (200, 200, 200), 1)


# ─────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────
def export_to_excel(csv_filepath: str, excel_filepath: str = "attendance.xlsx"):
    """
    Convert attendance CSV to a formatted Excel file.
    """
    if not OPENPYXL_AVAILABLE:
        print("[ERROR] openpyxl not installed. Install with: pip install openpyxl")
        return False

    if not os.path.isfile(csv_filepath):
        print(f"[ERROR] Attendance file '{csv_filepath}' not found.")
        return False

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Add headers
    headers = ["Name", "Time (HH:MM:SS)", "Date (YYYY-MM-DD)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Read CSV and add data
    row_num = 2
    with open(csv_filepath, "r", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = center_align
                cell.border = border
            row_num += 1

    # Auto-adjust column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    # Save
    wb.save(excel_filepath)
    print(f"[SUCCESS] Attendance exported to: {excel_filepath}")
    return True


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────
def main():
    print("=" * 50)
    print("  Face Recognition Attendance System")
    print("=" * 50)

    # 1. Load dataset
    known_encodings, known_names = load_dataset(DATASET_DIR)
    if not known_encodings:
        print("[ERROR] No encodings loaded. Check your dataset folder.")
        return

    # 2. Create screenshots directory
    os.makedirs(SCREENSHOTS_DIR, exist_ok=True)

    # 3. Ensure CSV has header row
    ensure_csv_header(ATTENDANCE_FILE)

    # FIX 1: Load today's already-marked names instead of wiping
    marked_today = load_todays_attendance(ATTENDANCE_FILE)
    if marked_today:
        print(f"[INFO] Session started. Already marked today: {', '.join(sorted(marked_today))}")
    else:
        print("[INFO] Session started. No one marked yet today.")

    # 4. Open webcam with fallback logic (FIX 9)
    cap = None
    for camera_idx in range(5):  # Try cameras 0-4
        cap = cv2.VideoCapture(camera_idx)
        if cap.isOpened():
            print(f"[INFO] Opened camera {camera_idx}")
            break
        cap.release()
    
    if cap is None or not cap.isOpened():
        print("[ERROR] Cannot open any webcam. Check camera connection.")
        return

    print("[INFO] Webcam started. Press 'Q' to quit, 'R' to reset.\n")

    tracker   = StabilityTracker(required_frames=CONFIRM_FRAMES)
    frame_num = 0
    
    # FIX 5: Non-blocking notification system
    notification_text = ""
    notification_expiry = 0  # timestamp

    while True:
        ret, frame = cap.read()
        if not ret:
            print("[WARN] Failed to grab frame. Retrying...")
            continue

        frame_num += 1

        # ── Process every Nth frame for performance ──
        if frame_num % PROCESS_EVERY_N == 0:

            # Resize to speed up recognition
            small_frame = cv2.resize(frame, (0, 0), fx=0.5, fy=0.5)
            rgb_small   = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

            # Detect faces
            face_locations = face_recognition.face_locations(rgb_small, model="hog")
            face_encodings = face_recognition.face_encodings(rgb_small, face_locations)

            if face_encodings:
                # FIX 3: Process ALL detected faces, not just the first one
                for face_enc, loc in zip(face_encodings, face_locations):
                    # Scale coordinates back to original frame size
                    top, right, bottom, left = [v * 2 for v in loc]

                    # Identify - returns (name, confidence)
                    raw_name, confidence = identify_face(face_enc, known_encodings, known_names, MATCH_THRESHOLD)

                    if raw_name == "Unknown":
                        tracker.update("Unknown")
                        draw_face_box(frame, top, right, bottom, left, "Unknown", False)
                    else:
                        # FIX 1: Normalize to lowercase
                        raw_name = raw_name.lower()
                        # Stability check
                        newly_confirmed, display_name = tracker.update(raw_name)

                        # Mark attendance on confirmation
                        if newly_confirmed and newly_confirmed not in marked_today:
                            # FIX 2: Save screenshot when attendance is marked
                            timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
                            screenshot_filename = f"{SCREENSHOTS_DIR}/{newly_confirmed}_{timestamp_str}.jpg"
                            # Use absolute path for screenshot storage
                            screenshot_abs_path = os.path.abspath(screenshot_filename)
                            cv2.imwrite(screenshot_abs_path, frame)
                            print(f"[INFO] Screenshot saved: {screenshot_abs_path}")
                            
                            # Save to CSV file
                            mark_attendance(newly_confirmed, ATTENDANCE_FILE, marked_today)
                            
                            # Save to database with screenshot path and confidence
                            save_attendance_to_database(newly_confirmed, screenshot_abs_path, confidence)
                            
                            # FIX 5: Non-blocking notification
                            notification_text = f"ATTENDANCE MARKED: {newly_confirmed}"
                            notification_expiry = time.time() + 2.0  # show for 2 seconds
                            tracker.reset()

                        # Draw box
                        confirmed_disp = tracker.confirmed_locked
                        draw_face_box(frame, top, right, bottom, left, display_name, confirmed_disp)

            else:
                # No face detected – hard reset streak for faster redetection (FIX 10)
                if not tracker.confirmed_locked:
                    tracker.streak = 0
                    tracker.candidate = None

        # FIX 6: Move display and key handling OUTSIDE the processing block
        # Always draw status bar
        draw_status_bar(frame, tracker, marked_today)

        # FIX 5: Draw notification if still active
        if time.time() < notification_expiry:
            cv2.putText(frame, notification_text,
                        (10, frame.shape[0] - 20),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 2)

        cv2.imshow("Attendance System", frame)

        # ── Single, centralised key handler ──
        key = cv2.waitKey(1) & 0xFF

        if key == ord('q'):
            print("\n[INFO] Quit requested by user.")
            break

        if key == ord('r'):
            # Reset: clear memory and tracker (but preserve CSV for session)
            print("[INFO] Resetting tracker...")
            marked_today.clear()
            ensure_csv_header(ATTENDANCE_FILE)  # reset CSV
            tracker.reset()
            notification_text = "RESET DONE - Ready for new marks"
            notification_expiry = time.time() + 1.0
            print("[INFO] Tracker reset. Attendance cleared.")

    cap.release()
    cv2.destroyAllWindows()
    print("[INFO] Session ended.")
    
    # Session summary
    all_people = set(known_names)
    absent_today = all_people - marked_today
    
    print("\n" + "=" * 50)
    print("  SESSION SUMMARY")
    print("=" * 50)
    print(f"Present : {', '.join(sorted(marked_today)) if marked_today else 'None'}")
    print(f"Absent  : {', '.join(sorted(absent_today)) if absent_today else 'None'}")
    print("=" * 50 + "\n")


if __name__ == "__main__":
    import sys

    print("\n" + "=" * 50)
    print("  Attendance System Menu")
    print("=" * 50)
    print("1. Run face recognition (default)")
    print("2. Export attendance to Excel")
    print("="*50)

    choice = input("\nSelect option (1/2) [default=1]: ").strip()

    if choice == "2":
        export_to_excel(ATTENDANCE_FILE)
    else:
        main()
